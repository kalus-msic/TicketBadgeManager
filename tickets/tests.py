from django.test import TestCase
from django.core.exceptions import ValidationError
from unittest.mock import patch, MagicMock
from tickets.utils.validators import validate_merge_file
from tickets.printing.profiles.tspl import TSPLProfile
from tickets.printing.backends.direct import DirectBackend
from tickets.printing.manager import PrintManager


class ValidateMergeFileTest(TestCase):

    def _make_file(self, name, size=100, content=b'col1;col2\nval1;val2'):
        f = MagicMock()
        f.name = name
        f.size = size
        f.read.return_value = content
        return f

    def test_accepts_csv(self):
        f = self._make_file('data.csv')
        self.assertTrue(validate_merge_file(f))

    def test_accepts_xlsx(self):
        f = self._make_file('data.xlsx', content=b'binarydata')
        # xlsx files don't need UTF-8 decode check
        self.assertTrue(validate_merge_file(f))

    def test_rejects_txt(self):
        f = self._make_file('data.txt')
        with self.assertRaises(ValidationError):
            validate_merge_file(f)

    def test_rejects_oversized(self):
        f = self._make_file('data.csv', size=11 * 1024 * 1024)
        with self.assertRaises(ValidationError):
            validate_merge_file(f)

    def test_rejects_none(self):
        with self.assertRaises(ValidationError):
            validate_merge_file(None)


import io
import pandas as pd
from tickets.utils.merge_utils import find_header_row, read_file_to_dataframe


class FindHeaderRowTest(TestCase):

    def test_finds_header_with_many_string_cols(self):
        rows = [
            ['Pořadatel', 'MSIC', None, None],
            ['Akce', 'InnoVerse', None, None],
            ['Příjmení', 'Jméno', 'Číslo objednávky', 'Číslo vstupenky', 'Kategorie'],
        ]
        self.assertEqual(find_header_row(rows), 2)

    def test_skips_rows_with_fewer_than_4_strings(self):
        rows = [
            ['Label', 'Value', None, None],
            ['Another', 'Row', None, None],
            ['Col1', 'Col2', 'Col3', 'Col4'],
        ]
        self.assertEqual(find_header_row(rows), 2)

    def test_raises_if_no_header_found(self):
        rows = [
            ['A', 'B', None],
            ['C', 'D', None],
        ]
        with self.assertRaises(ValueError):
            find_header_row(rows)

    def test_ignores_nan_strings(self):
        # When CSV is parsed with dtype=str, empty cells become 'nan'
        rows = [
            ['Pořadatel', 'MSIC', 'nan', 'nan'],
            ['Col1', 'Col2', 'Col3', 'Col4'],
        ]
        # First row: only 2 real strings ('nan' excluded), so header is row 1
        self.assertEqual(find_header_row(rows), 1)


class ReadFileToDfTest(TestCase):

    def _make_csv_bytes(self, content):
        return content.encode('utf-8-sig')

    def test_reads_csv_with_semicolon_and_metadata(self):
        content = (
            "Pořadatel;MSIC\n"
            "Akce;InnoVerse\n"
            "Číslo objednávky;Číslo vstupenky;Kategorie;Příjmení\n"
            "12345;5648915060501;Základní vstupné;Novák\n"
        )
        df = read_file_to_dataframe(self._make_csv_bytes(content), 'test.csv')
        self.assertIn('Číslo objednávky', df.columns)
        self.assertEqual(df.iloc[0]['Číslo objednávky'], '12345')

    def test_reads_csv_with_comma_fallback(self):
        content = (
            "Label,Value\n"
            "Col1,Col2,Col3,Col4\n"
            "a,b,c,d\n"
        )
        df = read_file_to_dataframe(self._make_csv_bytes(content), 'test.csv')
        self.assertIn('Col1', df.columns)

    def test_order_numbers_stay_as_strings(self):
        content = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Jméno;Příjmení;Kategorie\n"
            "34302271;Jana;Nová;Basic\n"
        )
        df = read_file_to_dataframe(self._make_csv_bytes(content), 'test.csv')
        self.assertEqual(df.iloc[0]['Číslo objednávky'], '34302271')
        self.assertNotIn('.0', df.iloc[0]['Číslo objednávky'])


import uuid
from datetime import date
from django.test import TestCase, Client
from django.urls import reverse
from django.core.cache import cache
from django.contrib.auth.models import User
from tickets.models import Event


class MergeExecuteViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='staff', password='pass', is_staff=True
        )
        self.client.login(username='staff', password='pass')
        self.event = Event.objects.create(name="Test Event", date=date(2026, 1, 1))

    def _make_pair(self):
        left = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Číslo vstupenky;Příjmení;Jméno\n"
            "111;VST001;Novák;Jan\n"
            "222;VST002;Dvořák;Eva\n"
        ).encode('utf-8-sig')
        right = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Název organizace;E-mail;Kategorie\n"
            "111;Firma s.r.o.;jan@example.com;Basic\n"
            "333;Jiná firma;jiri@example.com;VIP\n"
        ).encode('utf-8-sig')
        return left, right

    def _seed_cache(self, file1_bytes, file2_bytes):
        key = str(uuid.uuid4())
        cache.set(f'merge_{key}', {
            'file1_bytes': file1_bytes,
            'file1_name': 'odbaveni.csv',
            'file2_bytes': file2_bytes,
            'file2_name': 'transakce.csv',
        }, 60)
        return key

    def _reverse(self, name, **kwargs):
        return reverse(name, kwargs={'event_pk': self.event.pk, **kwargs})

    def test_get_redirects_to_merge_import(self):
        url = self._reverse('tickets:merge_execute')
        response = self.client.get(url)
        self.assertRedirects(response, self._reverse('tickets:merge_import'))

    def test_expired_cache_redirects_with_error(self):
        url = self._reverse('tickets:merge_execute')
        response = self.client.post(url, {
            'session_key': 'nonexistent',
            'join_column': 'Číslo objednávky',
            'action': 'import',
        })
        self.assertRedirects(response, self._reverse('tickets:merge_import'))

    def test_download_action_returns_csv(self):
        left, right = self._make_pair()
        key = self._seed_cache(left, right)
        url = self._reverse('tickets:merge_execute')
        response = self.client.post(url, {
            'session_key': key,
            'join_column': 'Číslo objednávky',
            'action': 'download',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('merged_goout.csv', response['Content-Disposition'])

    def test_download_does_not_delete_merge_cache(self):
        left, right = self._make_pair()
        key = self._seed_cache(left, right)
        self.client.post(self._reverse('tickets:merge_execute'), {
            'session_key': key,
            'join_column': 'Číslo objednávky',
            'action': 'download',
        })
        self.assertIsNotNone(cache.get(f'merge_{key}'))

    def test_import_action_creates_import_cache_and_redirects(self):
        left, right = self._make_pair()
        key = self._seed_cache(left, right)
        response = self.client.post(self._reverse('tickets:merge_execute'), {
            'session_key': key,
            'join_column': 'Číslo objednávky',
            'action': 'import',
        }, follow=False)
        self.assertEqual(response.status_code, 302)
        redirect_url = response['Location']
        self.assertIn('/import/mapping/', redirect_url)
        self.assertIn('session_key=', redirect_url)

    def test_import_deletes_merge_cache(self):
        left, right = self._make_pair()
        key = self._seed_cache(left, right)
        self.client.post(self._reverse('tickets:merge_execute'), {
            'session_key': key,
            'join_column': 'Číslo objednávky',
            'action': 'import',
        })
        self.assertIsNone(cache.get(f'merge_{key}'))

    def test_import_cache_has_correct_structure(self):
        left, right = self._make_pair()
        key = self._seed_cache(left, right)
        response = self.client.post(self._reverse('tickets:merge_execute'), {
            'session_key': key,
            'join_column': 'Číslo objednávky',
            'action': 'import',
        }, follow=False)
        redirect_url = response['Location']
        new_key = redirect_url.split('session_key=')[1]
        csv_data = cache.get(f'import_{new_key}')
        self.assertIsNotNone(csv_data)
        self.assertIn('fieldnames', csv_data)
        self.assertIn('rows', csv_data)
        self.assertIn('filename', csv_data)
        self.assertIn('delimiter', csv_data)
        # All row values must be strings
        for row in csv_data['rows']:
            for val in row.values():
                self.assertIsInstance(val, str)
        # No 'nan' values
        for row in csv_data['rows']:
            for val in row.values():
                self.assertNotEqual(val, 'nan')


class MergeImportViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='staff2', password='pass', is_staff=True
        )
        self.client.login(username='staff2', password='pass')
        self.event = Event.objects.create(name="Test Event", date=date(2026, 1, 1))

    def _reverse(self, name, **kwargs):
        return reverse(name, kwargs={'event_pk': self.event.pk, **kwargs})

    def test_get_renders_step1(self):
        response = self.client.get(self._reverse('tickets:merge_import'))
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('step', response.context or {})

    def test_post_valid_files_renders_step2(self):
        left = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Číslo vstupenky;Příjmení;Jméno\n"
            "111;VST001;Novák;Jan\n"
        ).encode('utf-8-sig')
        right = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Název organizace;E-mail;Kategorie\n"
            "111;Firma;jan@example.com;Basic\n"
        ).encode('utf-8-sig')

        left_file = io.BytesIO(left)
        left_file.name = 'odbaveni.csv'
        right_file = io.BytesIO(right)
        right_file.name = 'transakce.csv'

        response = self.client.post(
            self._reverse('tickets:merge_import'),
            {'file1': left_file, 'file2': right_file},
        )
        self.assertEqual(response.status_code, 200)
        ctx = response.context
        self.assertEqual(ctx['step'], 2)
        self.assertIn('Číslo objednávky', ctx['common_columns'])
        self.assertEqual(ctx['suggested_join_column'], 'Číslo objednávky')
        self.assertIn('session_key', ctx)

    def test_post_caches_file_bytes(self):
        left = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Číslo vstupenky;Příjmení;Jméno\n"
            "111;VST001;Novák;Jan\n"
        ).encode('utf-8-sig')
        right = (
            "Pořadatel;MSIC\n"
            "Číslo objednávky;Název organizace;E-mail;Kategorie\n"
            "111;Firma;jan@example.com;Basic\n"
        ).encode('utf-8-sig')

        left_file = io.BytesIO(left)
        left_file.name = 'odbaveni.csv'
        right_file = io.BytesIO(right)
        right_file.name = 'transakce.csv'

        response = self.client.post(
            self._reverse('tickets:merge_import'),
            {'file1': left_file, 'file2': right_file},
        )
        session_key = response.context['session_key']
        cached = cache.get(f'merge_{session_key}')
        self.assertIsNotNone(cached)
        self.assertIn('file1_bytes', cached)
        self.assertIn('file2_bytes', cached)

    def test_post_no_common_columns_redirects_with_error(self):
        left = (
            "Pořadatel;MSIC\n"
            "UniqueA;UniqueB;UniqueC;UniqueD\n"
            "a;b;c;d\n"
        ).encode('utf-8-sig')
        right = (
            "Pořadatel;MSIC\n"
            "DifferentA;DifferentB;DifferentC;DifferentD\n"
            "1;2;3;4\n"
        ).encode('utf-8-sig')

        left_file = io.BytesIO(left)
        left_file.name = 'file1.csv'
        right_file = io.BytesIO(right)
        right_file.name = 'file2.csv'

        response = self.client.post(
            self._reverse('tickets:merge_import'),
            {'file1': left_file, 'file2': right_file},
            follow=True,
        )
        self.assertRedirects(response, self._reverse('tickets:merge_import'))
        messages_list = list(response.context['messages'])
        self.assertTrue(any('common' in str(m).lower() or 'No common' in str(m) for m in messages_list))


class ImportMappingGetBranchTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='staff3', password='pass', is_staff=True
        )
        self.client.login(username='staff3', password='pass')
        self.event = Event.objects.create(name="Test Event", date=date(2026, 1, 1))

    def _reverse(self, name, **kwargs):
        return reverse(name, kwargs={'event_pk': self.event.pk, **kwargs})

    def _seed_import_cache(self):
        key = str(uuid.uuid4())
        cache.set(f'import_{key}', {
            'fieldnames': ['Číslo objednávky', 'Číslo vstupenky', 'Název organizace', 'Jméno'],
            'rows': [
                {'Číslo objednávky': '111', 'Číslo vstupenky': 'VST001', 'Název organizace': 'Firma', 'Jméno': 'Jan'},
            ],
            'filename': 'merged_goout.csv',
            'delimiter': ',',
        }, 60)
        return key

    def test_get_with_valid_session_key_renders_mapping(self):
        key = self._seed_import_cache()
        response = self.client.get(
            self._reverse('tickets:import_mapping') + f'?session_key={key}'
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'tickets/import_mapping.html')
        ctx = response.context
        self.assertEqual(len(ctx['csv_columns']), 4)
        self.assertEqual(ctx['total_rows'], 1)
        self.assertEqual(ctx['session_key'], key)
        self.assertEqual(ctx['delimiter'], ',')
        self.assertIn('delimiter_name', ctx)
        self.assertIn('detected_profile', ctx)
        self.assertIn('profile_name', ctx)

    def test_get_with_expired_key_redirects_with_error(self):
        response = self.client.get(
            self._reverse('tickets:import_mapping') + '?session_key=expired-key',
            follow=True,
        )
        self.assertRedirects(response, self._reverse('tickets:merge_import'))
        msgs = list(response.context['messages'])
        self.assertTrue(any('expired' in str(m).lower() for m in msgs))


from datetime import date
from tickets.models import Event, Ticket, Log


class EventModelTest(TestCase):
    def test_create_event(self):
        event = Event.objects.create(
            name="Test Conference",
            date=date(2026, 6, 15),
            description="A test event"
        )
        self.assertEqual(str(event), "Test Conference (2026-06-15)")
        self.assertEqual(event.status, 'active')
        self.assertEqual(event.printer_1_name, 'TDP-2251')
        self.assertEqual(event.printer_2_name, 'TDP-2252')
        self.assertTrue(event.auto_print_on_scan)

    def test_ticket_event_fk(self):
        event = Event.objects.create(name="Test", date=date(2026, 1, 1))
        ticket = Ticket.objects.create(
            qr_code="EVT-001", name="Test User", event=event
        )
        self.assertEqual(ticket.event, event)

    def test_log_event_fk(self):
        event = Event.objects.create(name="Test", date=date(2026, 1, 1))
        log = Log.objects.create(
            event_type='SYSTEM', message='Test log', event=event
        )
        self.assertEqual(log.event, event)

    def test_event_ordering_newest_first(self):
        e1 = Event.objects.create(name="Old", date=date(2025, 1, 1))
        e2 = Event.objects.create(name="New", date=date(2026, 1, 1))
        events = list(Event.objects.all())
        self.assertEqual(events[0], e2)
        self.assertEqual(events[1], e1)


import io as _io
from openpyxl import load_workbook as _load_workbook


class ExportXlsxViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='staff_xlsx', password='pass', is_staff=True
        )
        self.client.login(username='staff_xlsx', password='pass')
        self.event = Event.objects.create(name="Test Event", date=date(2026, 1, 1))

    def _reverse(self, name, **kwargs):
        return reverse(name, kwargs={'event_pk': self.event.pk, **kwargs})

    def _make_ticket(self):
        return Ticket.objects.create(
            qr_code='TEST001',
            name='Jana Nová',
            company_name='Firma s.r.o.',
            status='VALID',
            event=self.event,
        )

    def test_returns_200_with_xlsx_content_type(self):
        response = self.client.get(self._reverse('tickets:export_tickets_xlsx'))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type']
        )

    def test_content_disposition_is_attachment(self):
        response = self.client.get(self._reverse('tickets:export_tickets_xlsx'))
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])

    def test_response_is_valid_xlsx_with_correct_data(self):
        self._make_ticket()
        response = self.client.get(self._reverse('tickets:export_tickets_xlsx'))
        wb = _load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(1, 1).value, 'QR Code')
        self.assertEqual(ws.cell(1, 2).value, 'Name')
        self.assertEqual(ws.cell(2, 1).value, 'TEST001')
        self.assertEqual(ws.cell(2, 2).value, 'Jana Nová')

    def test_csv_export_still_works(self):
        response = self.client.get(self._reverse('tickets:export_tickets_csv'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])


class EventCRUDTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('staff', password='pass', is_staff=True)
        self.client.login(username='staff', password='pass')
        self.event = Event.objects.create(name="Test Event", date=date(2026, 6, 15))

    def test_event_list_shows_active_events(self):
        response = self.client.get('/events/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Test Event")

    def test_event_create(self):
        response = self.client.post('/events/create/', {
            'name': 'New Event', 'date': '2026-07-01', 'status': 'active'
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Event.objects.filter(name='New Event').exists())

    def test_event_edit(self):
        response = self.client.post(f'/events/{self.event.pk}/edit/', {
            'name': 'Updated', 'date': '2026-06-15', 'status': 'active',
            'printer_1_name': 'TDP-2251', 'printer_2_name': 'TDP-2252',
        })
        self.assertEqual(response.status_code, 302)
        self.event.refresh_from_db()
        self.assertEqual(self.event.name, 'Updated')

    def test_event_delete(self):
        response = self.client.post(f'/events/{self.event.pk}/delete/')
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Event.objects.filter(pk=self.event.pk).exists())

    def test_index_redirects_to_event_list(self):
        response = self.client.get('/')
        self.assertRedirects(response, '/events/')


class EventsContextProcessorTest(TestCase):
    def setUp(self):
        from django.test import RequestFactory
        self.factory = RequestFactory()
        self.event = Event.objects.create(name="Active", date=date(2026, 6, 1), status='active')
        Event.objects.create(name="Archived", date=date(2025, 1, 1), status='archived')

    def test_events_list_contains_only_active(self):
        from tickets.context_processors import events_context
        request = self.factory.get('/events/')
        request.resolver_match = None
        ctx = events_context(request)
        self.assertEqual(len(ctx['events_list']), 1)
        self.assertEqual(ctx['events_list'][0].name, "Active")
        self.assertIsNone(ctx['active_event'])

    def test_active_event_resolved_from_kwargs(self):
        from tickets.context_processors import events_context
        from unittest.mock import Mock
        request = self.factory.get(f'/events/{self.event.pk}/tickets/')
        request.resolver_match = Mock()
        request.resolver_match.kwargs = {'event_pk': self.event.pk}
        ctx = events_context(request)
        self.assertEqual(ctx['active_event'], self.event)


class TSPLProfileTest(TestCase):
    def test_generate_returns_bytes(self):
        profile = TSPLProfile()
        result = profile.generate({
            'name': 'Jan Novak',
            'company_name': 'MSIC',
            'qr_code': 'TEST123',
            'event_name': 'Test Event'
        })
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_generate_contains_tspl_commands(self):
        profile = TSPLProfile()
        result = profile.generate({
            'name': 'Jan Novak',
            'company_name': '',
            'qr_code': 'TEST456',
            'event_name': ''
        })
        # TSPL bitmap command should be present
        self.assertIn(b'BITMAP', result)

    def test_generate_image_returns_pil_image(self):
        from PIL import Image
        profile = TSPLProfile()
        result = profile.generate_image({
            'name': 'Jan Novak',
            'company_name': 'MSIC',
            'qr_code': 'TEST789',
            'event_name': 'Test'
        })
        self.assertIsInstance(result, Image.Image)
        # Image should be rotated (height > width after 90deg rotation)
        self.assertGreater(result.height, result.width)

    def test_generate_test_page(self):
        profile = TSPLProfile()
        result = profile.generate_test_page()
        self.assertIsInstance(result, bytes)
        self.assertIn(b'BITMAP', result)


from tickets.services.ticket_service import TicketService


class TicketServiceEventTest(TestCase):
    def setUp(self):
        self.event = Event.objects.create(name="Test", date=date(2026, 1, 1))
        self.other_event = Event.objects.create(name="Other", date=date(2026, 2, 1))
        Ticket.objects.create(qr_code="TICKET-ALICE-001", name="Alice", event=self.event)
        Ticket.objects.create(qr_code="TICKET-BOB-002", name="Bob", event=self.other_event)

    def test_get_statistics_scoped_to_event(self):
        stats = TicketService.get_statistics(self.event)
        self.assertEqual(stats['total'], 1)

    def test_search_tickets_scoped_to_event(self):
        results = TicketService.search_tickets(search_query='', event=self.event)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].name, 'Alice')

    def test_verify_ticket_creates_log_with_event(self):
        success, msg, ticket = TicketService.verify_ticket("TICKET-ALICE-001", self.event)
        self.assertTrue(success)
        from tickets.models import Log
        log = Log.objects.filter(event=self.event, event_type='CHECKIN').first()
        self.assertIsNotNone(log)


class DirectBackendTest(TestCase):
    @patch('tickets.printing.backends.direct.platform')
    def test_is_available_on_non_windows(self, mock_platform):
        mock_platform.system.return_value = 'Darwin'
        backend = DirectBackend()
        self.assertFalse(backend.is_available())

    @patch('tickets.printing.backends.direct.platform')
    def test_print_image_returns_printed_status(self, mock_platform):
        mock_platform.system.return_value = 'Windows'
        backend = DirectBackend()
        backend._tsclibrary = MagicMock()
        backend._printers_available = ['TDP-2251']
        from PIL import Image
        test_img = Image.new('L', (960, 580), 'white')
        result = backend.print_image(test_img, 'TDP-2251')
        self.assertEqual(result['status'], 'printed')
        backend._tsclibrary.openportW.assert_called_once_with('TDP-2251')
        backend._tsclibrary.sendcommandW.assert_any_call('CLS')
        backend._tsclibrary.printlabelW.assert_called_once_with('1', '1')
        backend._tsclibrary.closeport.assert_called_once()

    @patch('tickets.printing.backends.direct.platform')
    def test_print_image_fails_for_unknown_printer(self, mock_platform):
        mock_platform.system.return_value = 'Windows'
        backend = DirectBackend()
        backend._tsclibrary = MagicMock()
        backend._printers_available = ['TDP-2251']
        from PIL import Image
        test_img = Image.new('L', (960, 580), 'white')
        result = backend.print_image(test_img, 'UNKNOWN')
        self.assertEqual(result['status'], 'error')


class PrintManagerTest(TestCase):
    def setUp(self):
        from datetime import date
        self.event = Event.objects.create(
            name="Test Event", date=date(2026, 1, 1), print_backend="direct"
        )

    def test_get_printer_name_queue_1(self):
        pm = PrintManager(self.event)
        self.assertEqual(pm.get_printer_name("1"), "TDP-2251")

    def test_get_printer_name_queue_2(self):
        pm = PrintManager(self.event)
        self.assertEqual(pm.get_printer_name("2"), "TDP-2252")

    def test_print_direct_generates_image_and_dispatches(self):
        self.event.print_backend = "direct"
        self.event.save()
        pm = PrintManager(self.event)

        ticket_data = {
            'name': 'Jan Novak',
            'company_name': 'MSIC',
            'qr_code': 'QR123',
            'event_name': 'Test'
        }

        with patch.object(pm, '_backend') as mock_backend:
            mock_backend.print_image.return_value = {'status': 'printed'}
            result = pm.print_ticket(ticket_data, printer_queue="1")

        self.assertEqual(result['status'], 'printed')
        mock_backend.print_image.assert_called_once()
        from PIL import Image
        call_args = mock_backend.print_image.call_args
        self.assertIsInstance(call_args[0][0], Image.Image)
        self.assertEqual(call_args[0][1], "TDP-2251")

    def test_print_webusb_returns_print_required(self):
        self.event.print_backend = "webusb"
        self.event.save()
        pm = PrintManager(self.event)

        ticket_data = {
            'name': 'Jan Novak',
            'company_name': '',
            'qr_code': 'QR456',
            'event_name': ''
        }

        result = pm.print_ticket(ticket_data, printer_queue="1")
        self.assertEqual(result['status'], 'print_required')
        self.assertEqual(result['backend'], 'webusb')
        self.assertIn('data', result)

    def test_get_printer_name_falls_back_when_empty(self):
        self.event.printer_1_name = ""
        self.event.save()
        pm = PrintManager(self.event)
        self.assertEqual(pm.get_printer_name("1"), "TDP-2251")


class ScannerPrintBackendTest(TestCase):
    def setUp(self):
        from datetime import date
        self.event = Event.objects.create(
            name="Test Event", date=date(2026, 1, 1),
            print_backend="webusb"
        )
        self.ticket = Ticket.objects.create(
            qr_code="SCAN001", name="Test User",
            company_name="Test Co", event=self.event
        )

    def test_verify_webusb_returns_print_required(self):
        self.client.force_login(self._create_staff_user())
        response = self.client.post(
            f'/events/{self.event.pk}/verify/',
            {'qr_code': 'SCAN001', 'print': 'true', 'printer_queue': '1'}
        )
        data = response.json()
        self.assertTrue(data['success'])
        self.assertEqual(data.get('print_backend'), 'webusb')
        self.assertIn('print_data', data)

    def test_verify_direct_returns_print_result(self):
        self.event.print_backend = "direct"
        self.event.save()
        self.client.force_login(self._create_staff_user())
        response = self.client.post(
            f'/events/{self.event.pk}/verify/',
            {'qr_code': 'SCAN001', 'print': 'true', 'printer_queue': '1'}
        )
        data = response.json()
        self.assertTrue(data['success'])
        # On non-Windows, direct backend will return error but won't have print_backend key
        self.assertNotIn('print_backend', data)

    def _create_staff_user(self):
        from django.contrib.auth.models import User
        return User.objects.create_user(
            'staff', 'staff@test.com', 'pass', is_staff=True
        )
